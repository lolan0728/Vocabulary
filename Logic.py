from cProfile import label
import os
import random
import time
import datetime
import math
import pickle
import re
import configparser
import openpyxl
import calendar
import Outputer as outputer
import numpy as np
import matplotlib.pyplot as mp
from Vocabulary import Vocabulary, Vocabulary_Output, Word, WordType, DailyPerformance

class Logic:
    
    vocabListFile = '.\\data\\VocabList.pkl'
    currNewVocabsFile = '.\\data\\CurrNewVocab.pkl'
    currRevVocabsFile = '.\\data\\CurrRevVocab.pkl'
    uniVocabularyFile = '.\\data\\UniVocabulary.pkl'
    dailyWordsTableFile = '.\\data\\DailyWordsTable.pkl'
    configFile = '.\\settings\\settings.conf'

    def __init__(self):
        self.vocabList = {}
        # self.currVocabs = {}
        self.currNewVocabs = {}
        self.currRevVocabs = {}
        self.uniVocabulary = {}
        self.dailyWordsTable = {}
        self.dPerformance = {}
        self.mPerformance = {}
        self.wSummary = {}
        self.today = datetime.date.today().strftime("%Y/%m/%d")
        self.windows = {}
        self.readConf()

    def readConf(self):
        conf= configparser.ConfigParser()
        if os.path.exists(self.configFile):
            conf.read(self.configFile)
            self.name = conf.get("user", "name")

    def setCurrNewVocabs(self, vocabs):
        self.currNewVocabs = {}
        for name in vocabs:
            if name in self.vocabList.keys():
                self.currNewVocabs[name] = self.vocabList[name]

    def setCurrRevVocabs(self, vocabs):
        self.currRevVocabs = {}
        for name in vocabs:
            if name in self.vocabList.keys():
                self.currRevVocabs[name] = self.vocabList[name]
    
    # Make {word:newData} dictionary
    def getRevWordsDateList(self, today):
        self.revWordsDateList = {}
        for v in self.currRevVocabs.values():
            for v1 in v.words:
                v1 = self.uniVocabulary[v1]
                newDate = '' if ((v1.newDate is None) or 
                                    (v1.newDate == today)) else v1.newDate
                if v1.word not in self.revWordsDateList:
                    self.revWordsDateList[v1.word] = newDate

    def getNewWordsList(self, today):
        self.newWordsList = {}
        for v in self.currNewVocabs.values():
            for v1 in v.words:
                v1 = self.uniVocabulary[v1]
                newDate = '' if ((v1.newDate is None) or 
                                    (v1.newDate == today)) else v1.newDate
                if v1.word not in self.newWordsList:
                    if newDate == '':
                        self.newWordsList[v1.word] = newDate
    
    # Pick words from list by random with condition
    def getRevWordsBydate(self, dates, exc, quantity=None, isRandom=True):
        ls = [w for w in self.revWordsDateList if (self.revWordsDateList[w] 
                                                in dates and w not in exc)]
        if quantity is None:
            return ls
        quantity = quantity if len(ls) >= quantity else len(ls)
        if isRandom:
            return random.sample(ls, quantity)
        else:
            return ls[:quantity]

    def getNewWords(self, exc, quantity=None, isRandom=True):
        ls = [w for w in self.newWordsList if (w not in exc)]
        if quantity is None:
            return ls
        quantity = quantity if len(ls) >= quantity else len(ls)
        if isRandom:
            return random.sample(ls, quantity)
        else:
            return ls[:quantity]

    # Get all dates and sort them in reverse order
    def getRevDateList(self, today):
        ls = list(set([d for d in self.revWordsDateList.values() if d!='' and d<today])) 
        ls.sort(reverse=True) 
        return ls
    
    # Get all the information of the word
    def getWordInfo(self, w, wtype=WordType.OTHERS, isCloneReview=False):
        word = self.uniVocabulary[w].clone(isCloneReview)
        word.type = wtype
        return word

    # As the name
    def makeWordsTable(self, newEXC=None, reviewEXC=None, newNum=3,  
                                reviewNum=17, isRandom=True, today=None):

        today = self.today if today is None else today
        newEXC = [] if newEXC is None else [w.word for w in newEXC]
        reviewEXC = [] if reviewEXC is None else [w.word for w in reviewEXC]
        # allEXC = newEXC + reviewEXC

        newNum = newNum if newNum > 0 else 0
        reviewNum = reviewNum if reviewNum > 0 else 0
        
        wordsTable = {}
        self.getNewWordsList(today)
        self.getRevWordsDateList(today)
        datelist = self.getRevDateList(today)
        
        # New words
        newWords = self.getNewWords(newEXC, newNum, isRandom)
        
        # today-1
        priorityWords_1 = []
        if reviewNum > 0:
            if len(datelist) > 0:
                date = datelist[0]
                datelist = datelist[1:]
                priorityWords_1 = self.getRevWordsBydate([date], reviewEXC, reviewNum)
                reviewNum -= len(priorityWords_1)
        
        # today-1-1
        priorityWords_2 = []
        if reviewNum > 0:
            if len(datelist) > 0:
                date = datelist[0]
                datelist = datelist[1:]
                priorityWords_2 = self.getRevWordsBydate([date], reviewEXC, reviewNum)
                reviewNum -= len(priorityWords_2)

        # today-1-1-3
        # priorityWords_3 = []
        # if reviewNum > 0:
        #     if len(datelist) > 2:
        #         date = datelist[:2]
        #         datelist = datelist[3:]
        #         priorityWords_3 = self.getWordsBydate(date, allEXC)
        #         reviewNum -= len(priorityWords_3)

        otherWords = []
        if reviewNum > 0:
            if len(datelist) > 0:
                # otherWords = self.getWordsBydate(datelist, allEXC, reviewNum)
                others = self.getRevWordsBydate(datelist, reviewEXC)
                if len(others) > 0:
                    newdates = [self.uniVocabulary[w].newDate 
                                    if self.uniVocabulary[w].newDate else '' 
                                                                for w in others]
                    lastReviews = [self.uniVocabulary[w].getLastReviewDate()
                                    if self.uniVocabulary[w].getLastReviewDate() 
                                        else '' for w in others]
                    reviewTimes = [self.uniVocabulary[w].getReviewTimes() 
                                                                for w in others]
                    ind = np.lexsort((newdates, lastReviews, reviewTimes))
                    if len(ind) > reviewNum:
                        ind = ind[:reviewNum]
                    otherWords = [others[i] for i in ind]
        
        for ls in (newWords, priorityWords_1, priorityWords_2, otherWords):
            if ls == newWords:
                wType = WordType.NEW
            else:
                wType = WordType.REVIEW
            for w in ls:
                word = self.getWordInfo(w, wType)
                if word is not None:
                    wordsTable[word.word] = word
        return wordsTable
    
    def saveWordsTable(self, date, wordsTable):
        for w in self.uniVocabulary.values():
            if date == w.newDate:
                w.removeNewDate()
            if date in w.reviewDate:
                w.removeReviewDate(date)

        for w in wordsTable.values():
            if w.type == WordType.NEW:
                self.uniVocabulary[w.word].setNewDate(date)
            else:
                self.uniVocabulary[w.word].addReviewDate(date)
        self.dailyWordsTable[date] = wordsTable
        self.saveAll()

    def delWordFromAll(self, word):
        if word in self.uniVocabulary.keys():
            self.uniVocabulary.pop(word)
        for v in self.vocabList.values():
            v.deleteWord(word)
        self.saveAll()

    
    def saveDailyWordsTable(self, wordsTable, date=None):
        date = self.today if date is None else date
        self.dailyWordsTable[date] = wordsTable
        for w in self.dailyWordsTable.values():
            if w.type == WordType.NEW:
                self.uniVocabulary[w.word].setNewDate(date)
            else:
                self.uniVocabulary[w.word].addReviewDate(date)
        self.saveAll()
    
    def delVocabulary(self, vocab):
        if vocab in self.vocabList.keys():
            self.vocabList.pop(vocab)
        if vocab in self.currNewVocabs.keys():
            self.currNewVocabs.pop(vocab)
        if vocab in self.currRevVocabs.keys():
            self.currRevVocabs.pop(vocab)
        self.saveAll()

    def loadAll(self):
        self._loadVocabList()
        self._loadCurrNewVocabs()
        self._loadCurrRevVocabs()
        self._loadUniVocabulary()
        self._loadDailyWordsTable()

    def _loadVocabList(self):
        self._loadFromFile(self.vocabList, self.vocabListFile)

    def _loadCurrNewVocabs(self):
        self.currNewVocabs = {}
        self._loadFromFile(self.currNewVocabs, self.currNewVocabsFile)

    def _loadCurrRevVocabs(self):
        self.currRevVocabs = {}
        self._loadFromFile(self.currRevVocabs, self.currRevVocabsFile)
        
    def _loadUniVocabulary(self):
        self._loadFromFile(self.uniVocabulary, self.uniVocabularyFile)

    def fixWord(self, word, exp):
        self.uniVocabulary[word].explanation = exp

    def _loadDailyWordsTable(self):
        self._loadFromFile(self.dailyWordsTable, self.dailyWordsTableFile)

    def _loadFromFile(self, obj, path):
        data = self._loadDBFile(path)
        if data is not None:
            obj.update(data)

    # Load data from DataFile
    def _loadDBFile(self, path):
        if os.path.exists(path):
            with open(path, 'rb') as file:
                obj = pickle.loads(file.read())
            return obj
        return None

    def saveAll(self):
        self._saveVocabList()
        self._saveNewCurrVocabs()
        self._saveRevCurrVocabs()
        self._saveUniVocabulary()
        self._saveDailyWordsTable()

    def _saveVocabList(self):
        self._saveDBFile(self.vocabList, self.vocabListFile)
    
    def _saveNewCurrVocabs(self):
        self._saveDBFile(self.currNewVocabs, self.currNewVocabsFile)

    def _saveRevCurrVocabs(self):
        self._saveDBFile(self.currRevVocabs, self.currRevVocabsFile)

    def _saveUniVocabulary(self):
        self._saveDBFile(self.uniVocabulary, self.uniVocabularyFile)

    def _saveDailyWordsTable(self):
        self._saveDBFile(self.dailyWordsTable, self.dailyWordsTableFile)

    # Save data to DataFile
    def _saveDBFile(self, obj, path):
        with open(path, 'wb') as file:
            pickle.dump(obj, file, True)

    # Load Vocabulary from Excel file
    def loadExcel(self, path):
        workbook = openpyxl.load_workbook(path)
        for wsn in workbook.sheetnames:
            worksheet = workbook.get_sheet_by_name(wsn)
            vocabName = worksheet['A1'].value
            vocabdesc = worksheet['B1'].value
            v = Vocabulary(vocabName, vocabdesc)
            for row in list(worksheet.iter_rows())[2:]:
                word = row[0].value.strip()
                if len(word) > 0:
                    v.words.append(word)
                    exp = row[1].value
                    newdate = row[2].value
                    if row[2].value:
                        newdate = str(newdate).strip() \
                                        if str(newdate).strip() else None
                    if newdate is not None:
                        newdate = datetime.datetime \
                            .strptime(newdate, '%Y-%m-%d %H:%M:%S') \
                                                        .strftime('%Y/%m/%d')
                    if word not in self.uniVocabulary.keys():
                        self.uniVocabulary[word] = Word(word, exp, newdate)
                    else:
                        if newdate is not None:
                            if self.uniVocabulary[word].newDate is None or \
                                    self.uniVocabulary[word].newDate < newdate:
                                self.uniVocabulary[word].setNewDate(newdate)
            self.vocabList[vocabName] = v
        self.saveAll()

    def selfClean(self):
        self.uniVocabulary.pop(None)
        for v in self.vocabList.items():
            v.pop(None)
        self.saveAll()
    
    def exportAnalysis(self, vocabList, savepath):
        outVocabs = []
        for v in vocabList:
            vocab = self.vocabList[v]
            v_out = Vocabulary_Output(vocab.name, vocab.description)
            alldates = []
            for w in vocab.words:
                word = self.uniVocabulary[w]
                n = word.newDate
                if n is not None and len(n) > 0:
                    if n not in alldates:
                        alldates.append(n)
                reviewdates = word.reviewDate
                for r in reviewdates:
                    if r not in alldates:
                        alldates.append(r)
                v_out.words[w] = word
                alldates.sort()
                v_out.alldates = alldates
            outVocabs.append(v_out)
        outputer.Outputer().outputXlsx(outVocabs, savepath)

    def fuzzySearch(self, w):
        return [self.uniVocabulary[word].clone() 
                    for word in self.uniVocabulary.keys() 
                        if len(re.findall('^'+w, word, re.IGNORECASE)) > 0] \
                                                        if len(w) > 0 else []
    
    def preciseSearch(self, w):
        return self.uniVocabulary[w].clone() \
                    if w in self.uniVocabulary.keys() else None

    def getProgress(self, vocab):
        return len([w for w in self.vocabList[vocab].words \
                        if self.uniVocabulary[w].newDate is not None])
    
    def makeCalendar(self, year, month):
        return calendar.monthcalendar(year, month)

    def isVaildDate(self, date):
        try:
            time.strptime(date, "%Y/%m/%d")
            return True
        except:
            return False 
    
    def getPerformance(self):
        self._getDPerformance()
        self._getMPerformance()
        self._getWSummary()

    def _getDPerformance(self):
        self.dPerformance = {}
        for w in self.uniVocabulary.values():
            n = w.newDate
            if n is not None and len(n) > 0:
                if n not in self.dPerformance.keys():
                    dp = DailyPerformance(n, 1, 0)
                    self.dPerformance[n] = dp
                else:
                    dp = self.dPerformance[n]
                    dp.newWords += 1

            rews = w.reviewDate
            for r in rews:
                if r not in self.dPerformance.keys():
                    dp = DailyPerformance(r, 0, 1)
                    self.dPerformance[r] = dp
                else:
                    dp = self.dPerformance[r]
                    dp.reviews += 1
    
    def _getMPerformance(self):
        self.mPerformance = {}
        for v in self.dPerformance.values():
            ym = str(v.year) + '/' + str(v.month).rjust(2, '0')
            if ym in self.mPerformance.keys():
                self.mPerformance[ym][0] += v.newWords
                self.mPerformance[ym][1] += v.reviews
            else:
                self.mPerformance[ym] = [v.newWords, v.reviews]

    def _getWSummary(self):
        self.wSummary = {}
        for w in self.uniVocabulary.values():
            n = 1 if w.newDate is not None and len(w.newDate) > 0 else 0
            rews = len(w.reviewDate)
            cate = -1
            if n == 1:
                if rews <= 3:
                    cate = 0
                elif rews >= 4 and rews <= 5:
                    cate = 1
                # elif rews == 5:
                #     cate = 2
                elif rews > 5:
                    cate = 2
            self.wSummary[w.word] = cate
    
    def _getVSummary(self, desc, words):
        w_m1 = w_0 = w_1 = w_2 = 0
        for w in words:
            if self.wSummary[w] == -1:
                w_m1 += 1
            elif self.wSummary[w] == 0:
                w_0 += 1
            elif self.wSummary[w] == 1:
                w_1 += 1
            elif self.wSummary[w] == 2:
                w_2 += 1
            # elif self.wSummary[w] == 3:
            #     w_3 += 1
        return [desc, w_m1, w_0, w_1, w_2]
            
    def _getYMRange(self, yFrom, mFrom, yTo, mTo):
        ymRange = []
        if yFrom == yTo:
            for m in range(int(mFrom), int(mTo) + 1):
                ymRange.append(str(yFrom) + '/' + str(m).rjust(2, '0'))
        else:
            for y in range(int(yFrom), int(yTo) + 1):
                if y == int(yFrom):
                    for m in range(int(mFrom), 13):
                        ymRange.append(str(y) + '/' + str(m).rjust(2, '0'))
                elif int(yFrom)< y < int(yTo):
                    for m in range(1, 13):
                        ymRange.append(str(y) + '/' + str(m).rjust(2, '0'))
                elif y == int(yTo):
                    for m in range(1, int(mTo) + 1):
                        ymRange.append(str(y) + '/' + str(m).rjust(2, '0'))
        return ymRange

    def getMonthlyPerformance(self, yFrom, mFrom, yTo, mTo):
        ymRange = self._getYMRange(yFrom, mFrom, yTo, mTo)

        mPData = {}
        # mPData.append(['YearMonth', 'NewWords', 'Reviews', 'Workdays'])
        mPData['0000/00'] = ['年/月', '新单词数', '复习单词数', '学习日数']
        for ym in sorted(self.mPerformance.keys()):
            if ym in ymRange:
                # mPData.append([ym, self.mPerformance[ym][0], self.mPerformance[ym][1], 0])
                mPData[ym] = [ym, self.mPerformance[ym][0], self.mPerformance[ym][1], 0]
        
        for d in self.dailyWordsTable.keys():
            ym = d.split('/')[0]+ '/' + d.split('/')[1].rjust(2, '0')
            if ym in mPData:
                mPData[ym][3] += 1 
            
        return list(mPData.values())

    def getVocabularyData(self, yFrom, mFrom, yTo, mTo):
        ymRange = self._getYMRange(yFrom, mFrom, yTo, mTo)
        vData = []
        vData.append(['YearMonth', '词汇量'])
        start = 0
        for ym in sorted(self.mPerformance.keys()):
            vData.append([ym, start + self.mPerformance[ym][0]])
            start += self.mPerformance[ym][0]
        vData = [r for r in vData if r[0] == 'YearMonth' or r[0] in ymRange]
        return vData

    def getSummary(self):
        sData = []
        sData.append(['Description', '未掌握', '复习3次以下', '复习4-5次', '复习5次以上'])
        sData.append(self._getVSummary('总体情况', [w.word for w in self.uniVocabulary.values()]))
        for vocab in self.vocabList.values():
            sData.append(self._getVSummary(vocab.description, vocab.words))
        return sData

    # unused    
    def exportCharts(self, savepath, rng=None):
        self.getPerformance()
        if rng == None:
            ym = list(self.mPerformance.keys())
            ym.sort()
            f = ym[0].split('/')
            t = ym[-1].split('/')
            rng = [[f[0], f[1]], [t[0], t[1]]]
        # mPData = self.getMonthlyPerformance('2020', '12', '2021', '7')
        # past 10 months
        begin_year, begin_month = self.monthlimit(rng, -9)
        mPData = self.getMonthlyPerformance(begin_year, begin_month, rng[1][0], rng[1][1])
        # vData = self.getVocabularyData('2020', '12', '2021', '7')
        vData = self.getVocabularyData(begin_year, begin_month, rng[1][0], rng[1][1])
        sData = self.getSummary()
        outputer.Outputer().outputCharts(mPData, vData, sData, savepath)

    def monthlimit(self, ymRange, mlimit):
        begin_year,end_year = ymRange[0][0], ymRange[1][0]
        begin_month,end_month = ymRange[0][1], ymRange[1][1]
        
        if begin_month == end_year:
            months = int(end_month) - int(begin_month)
        else:
            months = (int(end_year) - int(begin_year))*12 + int(end_month) - int(begin_month)
        
        if months > mlimit:
           begin_year, begin_month = self.calMonth(end_year, end_month, mlimit)
        return begin_year, begin_month
    
    def calMonth(self, y, m, num):
        months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
        datamonth = int(str(y) + str(m).rjust(2,'0'))
        # datamonth = int(datamonth)
        num = int(num)
        year = datamonth // 100
        new_list = []
        s = math.ceil(abs(num) / 12)
        for i in range(int(-s), s + 1):
            new_list += [str(year + i) + x for x in months]
        new_list = [int(x) for x in new_list]
        result = str(new_list[new_list.index(datamonth) + num])
        return result[:4], result[4:]

    # For debug
    def printVocabulary(self, dicts):
        for v in dicts.values():
            if len(v.reviewDate) > 3:
                print('word :' + v.word)
                print('exp: ' + v.explanation)
                print('newDate: ' + str(v.newDate))
                print('reviewDate: ')
                for d in v.reviewDate:
                    print(d)

if __name__ == "__main__":
    #self tester
    t = Logic()
    t.loadAll()
    mPerformance = t.getMonthlyPerformance('2020', '12', '2021', '7')
    print(mPerformance)
    # for k, v in t.dPerformance.items():
    #     print("\nKey: " + k)
    #     print("year: " + str(v.year))
    #     print("month: " + str(v.month))
    #     print("newWords: " + str(v.newWords))
    #     print("reviews: " + str(v.reviews))
    # for k,v in mPData.items():
    #     print(k, v[0], v[1])
        
    # t.fixWord('well', '好的，井，水井')
    # t.printVocabulary(t.uniVocabulary)
    # t.saveAll()
    # for v in t.uniVocabulary.values():
    #     if isinstance(v.newDate, datetime.datetime):
    #         v.newDate = v.newDate.strftime("%Y/%m/%d")
    # t.saveAll()
    # savepath = r'C:\Users\eos\Desktop\test.xlsx'
    # t.exportAnalysis(t.vocabList, savepath)
    # for k,v in t.uniVocabulary.items():
    #     print(k + ': ' + v.explanation)
    # print('-'*20)
    # for v in t.vocabList.values():
    #     for w in v.words:
    #         print(w)
    # print('-'*20)
    # print(len(t.dailyWordsTable))




    
    

